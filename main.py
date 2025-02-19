import argparse
import openpyxl
from matplotlib.font_manager import FontProperties as font
import matplotlib.pyplot as plt

font1 = font(fname=r'fonts\NotoSansTC-Black.ttf')

def load_excel_to_data(filename: str) -> list:
    workbook = openpyxl.load_workbook(filename, read_only=True)
    worksheet = workbook.active
    data = [[]for _ in range(2)]
    for row in range(2):
        for col in range(1, worksheet.max_column):
            data[row].append(worksheet.cell(row=row+1, column=col+1).value)
    return data

def create_chart(filename: str, title: str, x_label: str, y_label: str, chart_type: str):
    data = load_excel_to_data(filename)
    if chart_type == 'line':
        plt.plot(data[0], data[1])
    elif chart_type == 'bar':
        plt.bar(data[0], data[1])
    plt.title(title, fontproperties=font1, fontsize=20)
    plt.xlabel(x_label, fontproperties=font1, fontsize=15)
    plt.ylabel(y_label, fontproperties=font1, fontsize=15)
    plt.show()

def main():
    parser = argparse.ArgumentParser(description='製作圖表')
    parser.add_argument('--excel', type=str, default= 'data.xlsx',  help='使用Excel檔案來生成圖表')
    parser.add_argument('--title', type=str, default= '圖表', help=' 圖表的標題')
    parser.add_argument('--x', type=str,default="X axis", help='X軸的標籤')
    parser.add_argument('--y', type=str,default='Y axis', help='Y軸的標籤')
    parser.add_argument('--type', type=str, default='line', help='圖表的類型')

    args = parser.parse_args()
    create_chart(args.excel, args.title, args.x, args.y, args.type)

if __name__ == '__main__':
    main()