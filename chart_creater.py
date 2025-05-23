from matplotlib.font_manager import FontProperties as font
import matplotlib.pyplot as plt
font1 = font(fname=r'fonts\NotoSansTC-Black.ttf')
import openpyxl

def load_excel_to_data(filename) -> list:
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    data = [[]for _ in range(ws.max_row)]
    x_list = []
    y_list = []
    result = []

    for i in range(1, ws.max_row+1):
        for j in range(1, ws.max_column+1):
            data[i-1].append(ws.cell(row=i, column=j).value)

    for i in range(2, ws.max_row+1):
        x_list.append(data[i-1][0])
        y_list.append(data[i-1][1])
    
    result.append(x_list)
    result.append(y_list)

    return result # result[0]為x座標 result[1]為y座標 

def analyze_data(data: list) -> list:
    # 數據計算
    info = []
    average = sum(data[1])/len(data[1])
    deviation = 0
    for i in data[1]:
        deviation += pow(i - average, 2)
    deviation /= len(data[1])
    info.append(average)
    info.append(deviation)
    x_max = max(data[0])
    x_min = min(data[0])
    y_max = max(data[1])
    y_min = min(data[1])
    info.append(x_min)
    info.append(x_max)
    info.append(y_min)
    info.append(y_max)
    return info # info[0]為平均數 info[1]為變異數 info[2]為x最小值 info[3]為x最大值 info[4]為y最小值 info[5]為y最大值


def create_chart(filename: str, title: str, x_label: str,
                  y_label: str, chart_type: str,
                  x_min: int, x_max: int, y_min: int, y_max: int, 
                  point_type_text: str, data_display: bool) -> plt.figure:
    data = load_excel_to_data(filename)
    point_type=''
    if point_type_text == '無':
        point_type = ''
    elif point_type_text == '圓形':
        point_type = 'o'
    elif point_type_text == '方形':
        point_type = 's'
    elif point_type_text == '三角形':
        point_type = '^'
    fig, ax = plt.subplots(figsize=(16/1.9, 9/1.9))
    #　選擇圖表類型
    if chart_type == '折線圖':
        ax.plot(data[0], data[1], marker=point_type)
    elif chart_type == '長條圖': # 長條圖沒有點樣式
        ax.bar(data[0], data[1])

    # 設置顯示範圍
    plt.xlim(x_min, x_max)
    plt.ylim(y_min, y_max)

    # 設置標題和標籤
    ax.set_title(title, fontproperties=font1, fontsize=20)
    ax.set_xlabel(x_label, fontproperties=font1, fontsize=15)
    ax.set_ylabel(y_label, fontproperties=font1, fontsize=15)
    
    # 數據顯示
    if data_display:
        for a, b in zip(data[0], data[1]):
            ax.text(a, b, b, ha='center', va='bottom', fontsize=10)
    info = analyze_data(data)
    # 平均數、離均差顯示
    ax.text(5.8, 0.5, f"平均數: {info[0]}\n變異數: {info[1]}", fontdict={'fontproperties': font1, 'fontsize': 15})
    return fig
